"""
Registry ingestion — builds backend/registry/data/registry.sqlite from RBI sources.

Run locally (needs openpyxl + pdfplumber, see requirements-ingest.txt):
    python -m backend.registry.ingest            # use cached downloads if present
    python -m backend.registry.ingest --fresh    # force re-download from RBI

Sources:
  1. RBI List of NBFCs & ARCs (XLSX) — name, classification, layer, deposit CoR,
     CIN, address, email. ~9,000 NBFCs + ~30 ARCs.
  2. RBI Scheduled + Non-Scheduled UCB lists (PDF) — head office address + pincode.
     ~1,900 cooperative banks. HEAD OFFICES ONLY by construction.
  3. Hardcoded: 11 Small Finance Banks (HQ city).

The SQLite file is committed to the repo so deploys need no DB setup.
"""

import argparse
import logging
import sqlite3
import sys
from pathlib import Path

import httpx

from backend.registry.geo import (
    city_coords, extract_city, extract_pin, jitter,
    state_from_cin, state_from_pin,
)

log = logging.getLogger("registry.ingest")

DATA_DIR = Path(__file__).parent / "data"
DB_PATH = DATA_DIR / "registry.sqlite"

RBI_NBFC_PAGE = "https://www.rbi.org.in/Scripts/BS_NBFCList.aspx"
RBI_NBFC_XLSX = ("https://rbidocs.rbi.org.in/rdocs/content/DOCs/"
                 "List_of_NBFCs_and_ARCs_registered_with_the_RBI.XLSX")
RBI_BANKS_PAGE = "https://www.rbi.org.in/commonperson/English/Scripts/banksinindia.aspx"
RBI_UCB_SCHED = "https://rbidocs.rbi.org.in/rdocs/content/pdfs/schedulecoop.pdf"
RBI_UCB_NONSCHED = "https://rbidocs.rbi.org.in/rdocs/content/pdfs/nonschedulecoop.pdf"
NSE_EQUITY_MASTER = "https://archives.nseindia.com/content/equities/EQUITY_L.csv"

_HEADERS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
                  "(KHTML, like Gecko) Chrome/125.0.0.0 Safari/537.36",
    "Accept-Language": "en-US,en;q=0.9",
}

SMALL_FINANCE_BANKS = [
    ("AU Small Finance Bank Ltd", "Jaipur", "Rajasthan"),
    ("Equitas Small Finance Bank Ltd", "Chennai", "Tamil Nadu"),
    ("Ujjivan Small Finance Bank Ltd", "Bengaluru", "Karnataka"),
    ("Suryoday Small Finance Bank Ltd", "Navi Mumbai", "Maharashtra"),
    ("ESAF Small Finance Bank Ltd", "Thrissur", "Kerala"),
    ("Utkarsh Small Finance Bank Ltd", "Varanasi", "Uttar Pradesh"),
    ("Jana Small Finance Bank Ltd", "Bengaluru", "Karnataka"),
    ("Capital Small Finance Bank Ltd", "Jalandhar", "Punjab"),
    ("North East Small Finance Bank Ltd", "Guwahati", "Assam"),
    ("Shivalik Small Finance Bank Ltd", "Noida", "Uttar Pradesh"),
    ("Unity Small Finance Bank Ltd", "New Delhi", "Delhi"),
]

SCHEMA = """
CREATE TABLE IF NOT EXISTS companies (
    id TEXT PRIMARY KEY,
    name TEXT NOT NULL,
    entity_type TEXT NOT NULL,        -- Bank | NBFC | ARC
    sub_type TEXT DEFAULT '',         -- UCB-Scheduled, UCB, SFB / ICC, MFI, P2P...
    layer TEXT DEFAULT '',            -- Base / Middle / Upper (NBFC scale)
    deposit_taking INTEGER DEFAULT 0,
    cin TEXT DEFAULT '',
    address TEXT DEFAULT '',
    city TEXT DEFAULT '',
    state TEXT DEFAULT '',
    pincode TEXT DEFAULT '',
    email TEXT DEFAULT '',
    rbi_region TEXT DEFAULT '',
    symbol TEXT DEFAULT '',          -- NSE trading symbol (listed companies)
    isin TEXT DEFAULT '',
    lat REAL, lng REAL,
    source TEXT NOT NULL,             -- rbi_nbfc_list / rbi_ucb_sched / ...
    ingested_at TEXT DEFAULT CURRENT_TIMESTAMP
);
CREATE INDEX IF NOT EXISTS idx_companies_city ON companies(city);
CREATE INDEX IF NOT EXISTS idx_companies_state ON companies(state);
CREATE INDEX IF NOT EXISTS idx_companies_type ON companies(entity_type);
CREATE INDEX IF NOT EXISTS idx_companies_pin ON companies(pincode);
"""


# ── Download (RBI needs a session cookie from a www page first) ──────────────

def _download(client: httpx.Client, warmup_page: str, url: str, dest: Path,
              magic: bytes) -> Path:
    if dest.exists() and dest.stat().st_size > 10_000 and not FRESH:
        log.info("cached: %s", dest.name)
        return dest
    client.get(warmup_page)
    r = client.get(url)
    if r.status_code != 200 or not r.content.startswith(magic):
        raise RuntimeError(
            f"download failed for {url}: status={r.status_code}, "
            f"head={r.content[:20]!r} (RBI anti-bot challenge — warmup page may have changed)"
        )
    dest.write_bytes(r.content)
    log.info("downloaded: %s (%d bytes)", dest.name, len(r.content))
    return dest


# ── Parsers ───────────────────────────────────────────────────────────────────

def parse_nbfc_xlsx(path: Path) -> list[dict]:
    import openpyxl
    wb = openpyxl.load_workbook(path, read_only=True)
    rows: list[dict] = []

    ws = wb["List of NBFCs"]
    for row in ws.iter_rows(min_row=3, values_only=True):
        # Sl.No | Name | Regional Office | Deposit CoR | Classification | CIN | Layer | Address | Email
        if not row or not row[1]:
            continue
        name = str(row[1]).strip()
        region = str(row[2] or "").strip()
        deposit = str(row[3] or "").strip().lower() == "yes"
        classification = str(row[4] or "").strip()
        cin = str(row[5] or "").strip()
        layer = str(row[6] or "").strip()
        address = str(row[7] or "").strip()
        email = str(row[8] or "").strip()
        pin = extract_pin(address)
        state = state_from_cin(cin) or state_from_pin(pin)
        city = extract_city(address, fallback=region)
        rows.append({
            "id": f"rbi_nbfc_{cin or name}",
            "name": name, "entity_type": "NBFC", "sub_type": classification,
            "layer": layer, "deposit_taking": int(deposit), "cin": cin,
            "address": address, "city": city, "state": state, "pincode": pin,
            "email": email, "rbi_region": region, "source": "rbi_nbfc_list",
        })

    ws = wb["ARCs"]
    for row in ws.iter_rows(min_row=3, values_only=True):
        # SR No | Name | Regional Office | CIN | Address | Email
        if not row or not row[1]:
            continue
        name = str(row[1]).strip()
        region = str(row[2] or "").strip()
        cin = str(row[3] or "").strip()
        address = str(row[4] or "").strip()
        email = str(row[5] or "").strip()
        pin = extract_pin(address)
        state = state_from_cin(cin) or state_from_pin(pin)
        rows.append({
            "id": f"rbi_arc_{cin or name}",
            "name": name, "entity_type": "ARC", "sub_type": "ARC",
            "layer": "", "deposit_taking": 0, "cin": cin,
            "address": address, "city": extract_city(address, fallback=region),
            "state": state, "pincode": pin, "email": email,
            "rbi_region": region, "source": "rbi_nbfc_list",
        })

    return rows


def parse_ucb_pdf(path: Path, scheduled: bool) -> list[dict]:
    import pdfplumber
    rows: list[dict] = []
    sub = "UCB-Scheduled" if scheduled else "UCB"
    src = "rbi_ucb_sched" if scheduled else "rbi_ucb_nonsched"
    with pdfplumber.open(path) as pdf:
        for page in pdf.pages:
            table = page.extract_table()
            if not table:
                continue
            for r in table:
                if not r or len(r) < 5 or not r[1]:
                    continue
                if str(r[0] or "").strip().lower() in ("sr. no.", "sr no", ""):
                    if not str(r[0] or "").strip().isdigit():
                        continue
                name = " ".join(str(r[1]).split())
                region = " ".join(str(r[2] or "").split()).title()
                address = " ".join(str(r[3] or "").split())
                pin = str(r[4] or "").strip()
                if not pin.isdigit():
                    pin = extract_pin(address)
                state = state_from_pin(pin)
                city = extract_city(address, fallback=region)
                rows.append({
                    "id": f"{src}_{name}_{pin}",
                    "name": name, "entity_type": "Bank", "sub_type": sub,
                    "layer": "", "deposit_taking": 1, "cin": "",
                    "address": address, "city": city, "state": state,
                    "pincode": pin, "email": "", "rbi_region": region,
                    "source": src,
                })
    return rows


def parse_nse_equity(path: Path) -> list[dict]:
    """All NSE-listed companies — symbol, official name, ISIN. The complete
    listed universe for Company Research (every sector, not just finance)."""
    import csv as _csv
    rows: list[dict] = []
    with open(path, newline="", encoding="utf-8-sig") as f:
        for r in _csv.DictReader(f):
            r = { (k or "").strip(): (v or "").strip() for k, v in r.items() }
            name = r.get("NAME OF COMPANY", "")
            symbol = r.get("SYMBOL", "")
            if not name or r.get("SERIES", "") not in ("EQ", "BE", "SM", "ST", ""):
                continue
            rows.append({
                "id": f"nse_{symbol}",
                "name": name, "entity_type": "Listed", "sub_type": "NSE-listed",
                "layer": "", "deposit_taking": 0, "cin": "",
                "address": "", "city": "", "state": "", "pincode": "",
                "email": "", "rbi_region": "",
                "symbol": symbol, "isin": r.get("ISIN NUMBER", ""),
                "source": "nse_equity_master",
            })
    return rows


def sfb_rows() -> list[dict]:
    out = []
    for name, city, state in SMALL_FINANCE_BANKS:
        out.append({
            "id": f"sfb_{name}",
            "name": name, "entity_type": "Bank", "sub_type": "SFB",
            "layer": "", "deposit_taking": 1, "cin": "", "address": f"{city}, {state}",
            "city": city, "state": state, "pincode": "", "email": "",
            "rbi_region": city, "source": "sfb_static",
        })
    return out


# ── Geocode (city centroid + deterministic jitter) ───────────────────────────

def assign_coords(rows: list[dict]) -> int:
    placed = 0
    for r in rows:
        r.setdefault("symbol", "")
        r.setdefault("isin", "")
        c = city_coords(r["city"]) or city_coords(r["rbi_region"])
        if c:
            r["lat"], r["lng"] = jitter(c[0], c[1], r["id"])
            placed += 1
        else:
            r["lat"], r["lng"] = None, None
    return placed


# ── Main ──────────────────────────────────────────────────────────────────────

FRESH = False


def main() -> None:
    global FRESH
    parser = argparse.ArgumentParser()
    parser.add_argument("--fresh", action="store_true", help="force re-download")
    FRESH = parser.parse_args().fresh

    logging.basicConfig(level=logging.INFO, format="%(levelname)s %(name)s: %(message)s")
    DATA_DIR.mkdir(parents=True, exist_ok=True)

    with httpx.Client(timeout=90, follow_redirects=True, headers=_HEADERS) as client:
        nbfc_x = _download(client, RBI_NBFC_PAGE, RBI_NBFC_XLSX,
                           DATA_DIR / "rbi_nbfc.xlsx", b"PK")
        ucb_s = _download(client, RBI_BANKS_PAGE, RBI_UCB_SCHED,
                          DATA_DIR / "ucb_scheduled.pdf", b"%PDF")
        ucb_n = _download(client, RBI_BANKS_PAGE, RBI_UCB_NONSCHED,
                          DATA_DIR / "ucb_nonscheduled.pdf", b"%PDF")
        nse_csv = _download(client, "https://www.nseindia.com", NSE_EQUITY_MASTER,
                            DATA_DIR / "nse_equity.csv", b"SYMBOL")

    rows = parse_nbfc_xlsx(nbfc_x)
    log.info("NBFC/ARC rows: %d", len(rows))
    ucb_rows = parse_ucb_pdf(ucb_s, scheduled=True) + parse_ucb_pdf(ucb_n, scheduled=False)
    log.info("UCB rows: %d", len(ucb_rows))
    listed = parse_nse_equity(nse_csv)
    log.info("NSE-listed rows: %d", len(listed))
    rows += ucb_rows + sfb_rows() + listed

    placed = assign_coords(rows)
    log.info("coords placed: %d / %d", placed, len(rows))

    if DB_PATH.exists():
        DB_PATH.unlink()
    con = sqlite3.connect(DB_PATH)
    con.executescript(SCHEMA)
    con.executemany(
        """INSERT OR REPLACE INTO companies
           (id, name, entity_type, sub_type, layer, deposit_taking, cin, address,
            city, state, pincode, email, rbi_region, symbol, isin, lat, lng, source)
           VALUES (:id, :name, :entity_type, :sub_type, :layer, :deposit_taking,
                   :cin, :address, :city, :state, :pincode, :email, :rbi_region,
                   :symbol, :isin, :lat, :lng, :source)""",
        rows,
    )
    con.commit()
    n = con.execute("SELECT COUNT(*) FROM companies").fetchone()[0]
    by_type = con.execute(
        "SELECT entity_type, COUNT(*) FROM companies GROUP BY entity_type"
    ).fetchall()
    con.close()
    log.info("registry.sqlite written: %d companies %s", n, by_type)


if __name__ == "__main__":
    main()
